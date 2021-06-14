#Tarea Programada 3
#Elaborado por: Felipe Obando y Sebastián Bermúdez.
#Fecha de creación: 01/06/2021
#Última modificación: 
#Versión: 3.9.2

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors,PatternFill,GradientFill
from archivos import *
from clases import *
from datetime import datetime
def reporteTotalidadLicencias():#Función para crear el reporte de todas las licencias
    """
    Función: Crear reporte de excel de totalidad de licencias.
    Entrada: N/A.
    Salida: N/A.
    """
    #a. Totalidad de licencias
    totalLicencias=lee("licencias")#trae todas la licencias
    listaLicencias=[]#lista para rellenar las filas.(DATOS)
    for i in totalLicencias:#crea la tupla de cada licencia que se va a guardar en la lista que se utiliza para las filas en el excel.
        if i.obtenerDonador():#esto está guardado como bool.
            donador="Si"
        else:
            donador="No"
        persona=(i.obtenerCedula(),i.obtenerNombreCompleto(),i.obtenerFechaNacimiento(),i.obtenerFechaExpedicion(),i.obtenerFechaVencimiento(),
        i.obtenerTipoLicencia(),i.obtenerTipoSangre(),donador,i.obtenerSede(),i.obtenerPuntaje())
        listaLicencias.append(persona)#guarda la fila (tupla) en las lista.

    fecha = datetime.now() #fecha y hora actual.
    fechaFormato = "Fecha y hora de creación: "+str(fecha.strftime("%d/%m/%Y, %H:%M:%S"))#para usar en el excel.
    
    wb = openpyxl.Workbook()#crea la instancia (archivo).
    hoja = wb.active#toma la instancia que crea wb y la utiliza como activa.
    hoja.title="Totalidad de licencias"#Asigna el nombre de la hoja.

    TITULO = Font(
                name='Arial',
                size=24,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000FF')
    FECHA = Font(
                name='Arial',
                size=14,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00000000')

    hoja.merge_cells('A1:F1')#Utiliza de la A1 a la G1 para el título
    hoja.merge_cells("A2:F2")#Utiliza de la A1 a la G1 para la fecha

    hoja["A1"] = "Reporte totalidad licencias"#TITULO 
    hoja["A1"].font=TITULO#formato del titulo, está arriba.

    hoja["A2"] = fechaFormato#Fecha y hora.
    hoja["A2"].font = FECHA#format de la fecha y hora.
    
    # Crea la fila del encabezado con los títulos
    hoja.append(('Cédula', 'Nombre completo', 
    'Fecha Nacimiento', 'Fecha Expedición',"Fecha Vencimiento",
    "Tipo licencia","Tipo Sangre","Donador","Sede","Puntaje"))#columna

    #Estas 4 lineas de código centran el texto de la A1 y A2.
    celdas = hoja["A1"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")
    celdas=hoja["A2"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")
    
    for licencia in listaLicencias:#saca los 
        hoja.append(licencia)
    try:

        wb.save("totalidadLicencias"+str(fecha.strftime("%d-%m-%Y"))+".xlsx")#guarda el archivo con el nombre.
    except:
        return False
    return True

def reporteTipoLicencia(tipo):
    """
    Función: Crear reporte de excel por tipo de licencia.
    Entrada: 
    -tipo(str): Tipo de licencia.
    Salida: N/A.
    """
    #b. Por tipo de licencia
    totalLicencias={"Licencias de conducir tipo A (motocicletas)":["Licencia A1","Licencia A2","Licencia A3"],
    "Licencias de conducir tipo B (automóviles y camiones)":["Licencia B1","Licencia B2 (camión pequeño)","Licencia B3 (camión pesado)","Licencia B4 (camión articulado)"],
    "Licencias de conducción tipo C (autobús y taxi)":["Licencia C1 (taxi)","Licencia C2 (autobús)"],
    "Licencias de conducir tipo D (tractores y maquinaria)":["Licencia D1","Licencia D2","Licencia D3"],
    "Licencias tipo E (universales)":["Licencia E1","Licencia E2"]}
    baseDatos=lee("licencias")
    listaLicencias=[]

    for licencia in baseDatos:
        if licencia.obtenerTipoLicencia() in totalLicencias[tipo]:
            listaLicencias.append((licencia.obtenerCedula(),licencia.obtenerNombreCompleto(),licencia.obtenerTipoLicencia()))

    fecha = datetime.now() #fecha y hora actual.
    fechaFormato = "Fecha y hora de creación: "+str(fecha.strftime("%d/%m/%Y, %H:%M:%S"))#para usar en el excel.
    
    wb = openpyxl.Workbook()#crea la instancia (archivo).
    hoja = wb.active#toma la instancia que crea wb y la utiliza como activa.
    hoja.title="Tipo de licencia"#Asigna el nombre de la hoja.

    TITULO = Font(
                name='Arial',
                size=18,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000FF')
    FECHA = Font(
                name='Arial',
                size=14,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00000000')

    hoja.merge_cells('A1:L1')#Utiliza de la A1 a la G1 para el título
    hoja.merge_cells("A2:F2")#Utiliza de la A1 a la G1 para la fecha

    hoja["A1"] = "Reporte "+ tipo#TITULO 
    hoja["A1"].font=TITULO#formato del titulo, está arriba.

    hoja["A2"] = fechaFormato#Fecha y hora.
    hoja["A2"].font = FECHA#format de la fecha y hora.
    
    # Crea la fila del encabezado con los títulos
    hoja.append(('Cédula', 'Nombre completo', "Tipo de licencia"))#columna

    #Estas 4 lineas de código centran el texto de la A1 y A2.
    celdas = hoja["A1"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")
    celdas=hoja["A2"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")

    for i in listaLicencias:
        hoja.append(i)
    try:
        paraNombreReporte=tipo.replace(" ","")
        wb.save("Reporte"+paraNombreReporte.capitalize()+".xlsx")#guarda el archivo con el nombre.
    except:
        return False
    return True


def reporteExamenSancion():
    """
    Función: Crear reporte de excel de las licencias con 0 como puntaje.
    Entrada: N/A.
    Salida: N/A.
    """
    #c.Examen por sanción
    baseDatos=lee("licencias")
    listaLicencias=[]

    for licencia in baseDatos:
        if licencia.obtenerPuntaje()<=6 and licencia.obtenerPuntaje()!=0:
            listaLicencias.append((licencia.obtenerCedula(),licencia.obtenerNombreCompleto(),licencia.obtenerTipoLicencia(),licencia.obtenerPuntaje()))

    fecha = datetime.now() #fecha y hora actual.
    fechaFormato = "Fecha y hora de creación: "+str(fecha.strftime("%d/%m/%Y, %H:%M:%S"))#para usar en el excel.
    
    wb = openpyxl.Workbook()#crea la instancia (archivo).
    hoja = wb.active#toma la instancia que crea wb y la utiliza como activa.
    hoja.title="Tipo de licencia"#Asigna el nombre de la hoja.

    TITULO = Font(
                name='Arial',
                size=18,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000FF')
    FECHA = Font(
                name='Arial',
                size=14,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00000000')

    hoja.merge_cells('A1:F1')#Utiliza de la A1 a la G1 para el título
    hoja.merge_cells("A2:F2")#Utiliza de la A1 a la G1 para la fecha

    hoja["A1"] = "Reporte examen por sanción"#TITULO 
    hoja["A1"].font=TITULO#formato del titulo, está arriba.

    hoja["A2"] = fechaFormato#Fecha y hora.
    hoja["A2"].font = FECHA#format de la fecha y hora.
    
    # Crea la fila del encabezado con los títulos
    hoja.append(('Cédula', 'Nombre completo', "Tipo de licencia","Puntaje"))#columna

    #Estas 4 lineas de código centran el texto de la A1 y A2.
    celdas = hoja["A1"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")
    celdas=hoja["A2"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")

    for i in listaLicencias:
        hoja.append(i)
    try:
        wb.save("ReporteExamenPorSancion"+".xlsx")#guarda el archivo con el nombre.
    except:
        return False
    return True

def reporteDonanteOrganos():
    baseDatos=lee("licencias")
    listaLicencias=[]
    for persona in baseDatos:
        if persona.obtenerDonador()==True:###
            listaLicencias.append((persona.obtenerCedula(),persona.obtenerNombreCompleto(),persona.obtenerTipoLicencia()))###

    fecha = datetime.now() #fecha y hora actual.
    fechaFormato = "Fecha y hora de creación: "+str(fecha.strftime("%d/%m/%Y, %H:%M:%S"))#para usar en el excel.

    wb = openpyxl.Workbook()#crea la instancia (archivo).
    hoja = wb.active#toma la instancia que crea wb y la utiliza como activa.
    hoja.title="Reporte de Donantes de Órganos"#Asigna el nombre de la hoja.###

    TITULO = Font(
                name='Arial',
                size=18,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000FF')
    FECHA = Font(
                name='Arial',
                size=14,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00000000')
    hoja.merge_cells('A1:F1')#Utiliza de la A1 a la G1 para el título
    hoja.merge_cells("A2:F2")#Utiliza de la A1 a la G1 para la fecha

    hoja["A1"] = "Donantes de Órganos"#TITULO ###
    hoja["A1"].font=TITULO#formato del titulo, está arriba.

    hoja["A2"] = fechaFormato#Fecha y hora.
    hoja["A2"].font = FECHA#format de la fecha y hora.

    hoja.append(('Cédula', 'Nombre completo', "Tipo de licencia"))#columna###

    celdas = hoja["A1"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")
    celdas=hoja["A2"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")

    for i in listaLicencias:
        hoja.append(i)
    try:
        wb.save("ReportePersonasDonantes"+".xlsx")#guarda el archivo con el nombre.###
    except:
        return False
    return True
def reporteLicenciaAnulada():
    baseDatos=lee("licencias")
    listaLicencias=[]
    for persona in baseDatos:
        if persona.obtenerPuntaje()==0:###
            if persona.obtenerDonador():
                donador="Sí"
            else:
                donador="No"
            listaLicencias.append((persona.obtenerCedula(),persona.obtenerNombreCompleto(),persona.obtenerFechaNacimiento(),persona.obtenerFechaExpedicion(),
            persona.obtenerFechaVencimiento(),persona.obtenerTipoLicencia(),persona.obtenerTipoSangre(),donador,persona.obtenerSede()))###

    fecha = datetime.now() #fecha y hora actual.
    fechaFormato = "Fecha y hora de creación: "+str(fecha.strftime("%d/%m/%Y, %H:%M:%S"))#para usar en el excel.

    wb = openpyxl.Workbook()#crea la instancia (archivo).
    hoja = wb.active#toma la instancia que crea wb y la utiliza como activa.
    hoja.title="Reporte de Personas Anuladas"#Asigna el nombre de la hoja.###

    TITULO = Font(
                name='Arial',
                size=18,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000FF')
    FECHA = Font(
                name='Arial',
                size=14,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00000000')
    hoja.merge_cells('A1:F1')#Utiliza de la A1 a la G1 para el título
    hoja.merge_cells("A2:F2")#Utiliza de la A1 a la G1 para la fecha

    hoja["A1"] = "Personas anuladas"#TITULO ###
    hoja["A1"].font=TITULO#formato del titulo, está arriba.

    hoja["A2"] = fechaFormato#Fecha y hora.
    hoja["A2"].font = FECHA#format de la fecha y hora.

    hoja.append(("Cédula","Nombre", "Fecha de Nacimiento", "Fecha de Expedición" ,"Fecha de Vencimiento" ,"Tipo de Licencia" ,"Tipo de Sangre" ,"Donador" ,"Sede"))#columna###

    celdas = hoja["A1"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")
    celdas=hoja["A2"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")

    for i in listaLicencias:
        hoja.append(i)
    try:
        wb.save("ReportePersonasAnuladas"+".xlsx")#guarda el archivo con el nombre.###
    except:
        return False
    return True
def reporteLicenciasPorSede(sede):
    baseDatos=lee("licencias")
    listaLicencias=[]
    for persona in baseDatos:
        if persona.obtenerSede()==sede:
            if persona.obtenerDonador():
                donador="Sí"
            else:
                donador="No"
            listaLicencias.append((persona.obtenerCedula(),persona.obtenerNombreCompleto(),persona.obtenerFechaNacimiento(),persona.obtenerFechaExpedicion(),
            persona.obtenerFechaVencimiento(),persona.obtenerTipoLicencia(),persona.obtenerTipoSangre(),donador,persona.obtenerSede(),str(persona.obtenerPuntaje())))###

    fecha = datetime.now() #fecha y hora actual.
    fechaFormato = "Fecha y hora de creación: "+str(fecha.strftime("%d/%m/%Y, %H:%M:%S"))#para usar en el excel.

    wb = openpyxl.Workbook()#crea la instancia (archivo).
    hoja = wb.active#toma la instancia que crea wb y la utiliza como activa.
    hoja.title="Reporte de Personas Por Sede"#Asigna el nombre de la hoja.###

    TITULO = Font(
                name='Arial',
                size=18,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000FF')
    FECHA = Font(
                name='Arial',
                size=14,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00000000')
    hoja.merge_cells('A1:F1')#Utiliza de la A1 a la G1 para el título
    hoja.merge_cells("A2:F2")#Utiliza de la A1 a la G1 para la fecha

    hoja["A1"] = "Licencias de "+sede#TITULO ###
    hoja["A1"].font=TITULO#formato del titulo, está arriba.

    hoja["A2"] = fechaFormato#Fecha y hora.
    hoja["A2"].font = FECHA#format de la fecha y hora.

    hoja.append(("Cédula","Nombre","Fecha de Nacimiento","Fecha de Expedición","Fecha de Vencimiento","Tipo de Licencia","Tipo de Sangre","Donador","Sede","Puntaje"))

    celdas = hoja["A1"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")
    celdas=hoja["A2"]
    celdas.alignment = Alignment(horizontal="center", vertical="center")
    for i in listaLicencias:
        hoja.append(i)
    try:
        paraNombre=sede.replace(",","")
        paraNombre2=paraNombre.replace(" ","")
        wb.save("ReportePersonasDe"+paraNombre2+".xlsx")#guarda el archivo con el nombre.###
    except:
        return False
    return True